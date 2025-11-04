#!/usr/bin/env python3
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill

# ================= RUTAS =================
BASE = Path("/Users/joseluisgiadanscastellanos/Library/CloudStorage/OneDrive-CEMEX/OneDrive_Cemex")
OUT_FILE = BASE / "Codigos/out/trafos_maestro_tabla.xlsx"

# ================= FUNCIONES =================
def calc_duval(row):
    """Calcula el diagnóstico según el Triángulo Duval 1 (CH4, C2H4, C2H2)."""
    ch4 = row.get("CH4", 0)
    c2h4 = row.get("C2H4", 0)
    c2h2 = row.get("C2H2", 0)

    total = ch4 + c2h4 + c2h2
    if total == 0:
        return 0, 0, 0, "Sin datos"

    # Porcentajes
    p_ch4 = (ch4 / total) * 100
    p_c2h4 = (c2h4 / total) * 100
    p_c2h2 = (c2h2 / total) * 100

    # Clasificación según zonas del triángulo Duval 1
    diag = "Indeterminado"

    # PD – Descargas Parciales
    if p_c2h2 < 4 and p_c2h4 < 10 and p_ch4 > 90:
        diag = "PD (Descargas Parciales)"

    # D1 – Descarga de baja energía
    elif 23 <= p_c2h2 <= 50 and 13 <= p_c2h4 <= 23:
        diag = "D1 (Descarga Baja Energía)"

    # D2 – Descarga de alta energía / Arco
    elif p_c2h2 > 50 and p_c2h4 < 20:
        diag = "D2 (Arco)"

    # T1 – Falla térmica < 300°C
    elif p_c2h4 < 20 and p_c2h2 < 4 and p_ch4 > 80:
        diag = "T1 (<300°C)"

    # T2 – Falla térmica 300–700°C
    elif 20 <= p_c2h4 <= 50 and p_c2h2 < 4 and 30 <= p_ch4 <= 80:
        diag = "T2 (300–700°C)"

    # T3 – Falla térmica >700°C
    elif p_c2h4 > 50 and p_c2h2 < 4 and p_ch4 < 20:
        diag = "T3 (>700°C)"

    # DT – Descarga + Térmica
    elif 20 < p_c2h2 < 50 and 20 < p_c2h4 < 40:
        diag = "DT (Discharge + Thermal)"

    return round(p_ch4, 2), round(p_c2h4, 2), round(p_c2h2, 2), diag

# ================= MAIN =================
def main():
    if not OUT_FILE.exists():
        print(f"❌ No encuentro el archivo: {OUT_FILE}")
        return

    # Leer hoja UltimaPorTrafo
    df = pd.read_excel(OUT_FILE, sheet_name="UltimaPorTrafo")

    # Asegurar columnas
    needed = ["CH4", "C2H4", "C2H2"]
    for col in needed:
        if col not in df.columns:
            print(f"❌ Falta columna {col}")
            return

    # Calcular diagnóstico Duval
    results = []
    for _, row in df.iterrows():
        p_ch4, p_c2h4, p_c2h2, diag = calc_duval(row)
        results.append({
            "Planta": row["Planta"],
            "Transformador": row["Transformador"],
            "Ubicacion": row["Ubicacion"],
            "Fecha de Muestra": row["Fecha de Muestra"],
            "%CH4": p_ch4,
            "%C2H4": p_c2h4,
            "%C2H2": p_c2h2,
            "Diagnóstico Duval": diag
        })

    duval_df = pd.DataFrame(results)

    # Guardar en hoja nueva
    with pd.ExcelWriter(OUT_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        duval_df.to_excel(writer, index=False, sheet_name="Diag_Duval")

    # ---- Formato tabla ----
    wb = load_workbook(OUT_FILE)
    ws = wb["Diag_Duval"]

    nrows, ncols = duval_df.shape
    table = Table(displayName="Tabla_Duval", ref=f"A1:H{nrows+1}")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # ---- Colorear diagnósticos ----
    col_diag = None
    for i, cell in enumerate(ws[1], start=1):
        if cell.value == "Diagnóstico Duval":
            col_diag = i
            break

    if col_diag:
        for r in range(2, ws.max_row + 1):
            val = str(ws.cell(r, col_diag).value or "")
            if "T1" in val:
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde
            elif "T2" in val:
                fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Amarillo
            elif "T3" in val or "D2" in val or "DT" in val:
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rojo
            elif "PD" in val:
                fill = PatternFill(start_color="9AD0F5", end_color="9AD0F5", fill_type="solid")  # Azul
            else:
                fill = None
            if fill:
                ws.cell(r, col_diag).fill = fill

    wb.save(OUT_FILE)
    print(f"✅ Hoja 'Diag_Duval' añadida con éxito a {OUT_FILE}")

if __name__ == "__main__":
    main()