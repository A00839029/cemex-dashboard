#!/usr/bin/env python3
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ================= RUTAS =================
BASE = Path("/Users/joseluisgiadanscastellanos/Library/CloudStorage/OneDrive-CEMEX/OneDrive_Cemex")
OUT_FILE = BASE / "Codigos/out/trafos_maestro_tabla.xlsx"

# ================= LIMITES IEC 60599 ================
LIMITS_IEC = {
    "H2": [(100, "Normal"), (700, "Preocupante"), (float("inf"), "Crítico")],
    "CH4": [(120, "Normal"), (1000, "Preocupante"), (float("inf"), "Crítico")],
    "C2H6": [(65, "Normal"), (1000, "Preocupante"), (float("inf"), "Crítico")],
    "C2H4": [(50, "Normal"), (1000, "Preocupante"), (float("inf"), "Crítico")],
    "C2H2": [(3, "Normal"), (50, "Preocupante"), (float("inf"), "Crítico")],
}

def clasificar_iec(valor, gas):
    try:
        v = float(valor)
    except:
        return "Normal"
    for lim, estado in LIMITS_IEC[gas]:
        if v <= lim:
            return estado
    return "Normal"

def diagnostico_iec(row):
    estados = []
    for gas in LIMITS_IEC.keys():
        if gas in row:
            estados.append(clasificar_iec(row[gas], gas))
    if "Crítico" in estados:
        return "Crítico"
    if "Preocupante" in estados:
        return "Preocupante"
    return "Normal"

def main():
    if not OUT_FILE.exists():
        print(f"❌ No se encontró {OUT_FILE}")
        return

    df = pd.read_excel(OUT_FILE, sheet_name="UltimaPorTrafo")
    df["Diagnóstico IEC"] = df.apply(diagnostico_iec, axis=1)

    cols = ["Planta", "Transformador", "Ubicacion", "Fecha de Muestra", "Diagnóstico IEC"]
    iec_df = df[cols].copy()

    wb = load_workbook(OUT_FILE)
    if "Diag_IEC" not in wb.sheetnames:
        wb.create_sheet("Diag_IEC")
    ws = wb["Diag_IEC"]

    # Limpiar contenido
    for row in ws["A2": f"Z{ws.max_row}"]:
        for cell in row:
            cell.value = None

    # Copiar datos
    for j, col in enumerate(iec_df.columns, start=1):
        ws.cell(1, j).value = col
    for i, row in enumerate(iec_df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j).value = val

    # Colorear Diagnóstico IEC
    colores = {"Normal": "C6EFCE", "Preocupante": "FFF2CC", "Crítico": "FFC7CE"}
    col_idx = iec_df.columns.get_loc("Diagnóstico IEC") + 1
    for i in range(2, ws.max_row + 1):
        val = ws.cell(i, col_idx).value
        if val in colores:
            ws.cell(i, col_idx).fill = PatternFill(start_color=colores[val], end_color=colores[val], fill_type="solid")

    wb.save(OUT_FILE)
    print(f"✅ Hoja 'Diag_IEC' creada y coloreada correctamente en {OUT_FILE}")

if __name__ == "__main__":
    main()