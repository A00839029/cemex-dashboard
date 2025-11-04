#!/usr/bin/env python3
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import pandas as pd
import unicodedata, re, sys

# ================= RUTAS BASE =================
BASE = Path("/Users/joseluisgiadanscastellanos/Library/CloudStorage/OneDrive-CEMEX/OneDrive_Cemex")
OUT_DIR = BASE / "Codigos/out"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_FILE = OUT_DIR / "trafos_maestro_tabla.xlsx"

# ============== CONFIG TABLA DGA ==============
SRC_FIRST_DATA_ROW = 16
SRC_START_COL = "AT"
SRC_LAST_COL = "BH"

# ================= HELPERS ====================
def _norm_str(s):
    if s is None:
        return ""
    return str(s).replace("\n", " ").strip()

def _key(s):
    s = "" if s is None else str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"\s+", " ", s).lower().strip()
    return s

def hoja_ruidosa(name: str) -> bool:
    SKIP = (
        "indice","índice","bdatos","pbianual","normas","cromatogra",
        "fisicas trafo","fisicoquimico","fisicoquímico","bpc","furan",
        "blank","ejemplo","example","plantilla","template","formato"
    )
    return any(tok in _key(name) for tok in SKIP)

def get(ws, addr):
    try:
        return ws[addr].value
    except Exception:
        return None

def _buscar_label_valor(ws, label, top=1, left=1, bottom=80, right=120):
    for r in range(top, bottom + 1):
        for c in range(left, right + 1):
            v = ws.cell(r, c).value
            if v and label.lower() in str(v).lower():
                val = ws.cell(r, c + 1).value if c + 1 <= right else None
                if val:
                    return _norm_str(val)
    return ""

def read_name_loc_from_sheet(ws):
    nombre = _norm_str(get(ws, "G9")) or _buscar_label_valor(ws, "Nombre")
    ubic = (
        _norm_str(get(ws, "H9"))
        or _norm_str(get(ws, "G11"))
        or _norm_str(get(ws, "C5"))
        or _buscar_label_valor(ws, "Ubicación")
    )
    return nombre, ubic

# ============== LECTURA ÍNDICE ===============
def leer_pares_indice_wb(wb):
    idx_name = None
    for n in wb.sheetnames:
        if "indice" in _key(n) or "índice" in _key(n):
            idx_name = n
            break
    if not idx_name:
        return set()

    ws = wb[idx_name]
    col_ubi = col_nom = header_row = None
    for r in range(1, 120):
        for c in range(1, 200):
            v = ws.cell(r, c).value
            if v is None:
                continue
            if col_ubi is None and "ubic" in _key(v):
                col_ubi = c
                header_row = r
            if col_nom is None and "nombre" in _key(v):
                col_nom = c
                header_row = r
        if col_ubi and col_nom:
            break

    pares = set()
    empty = 0
    for r in range(header_row + 1, header_row + 1 + 5000):
        nom = _norm_str(ws.cell(r, col_nom).value)
        ubi = _norm_str(ws.cell(r, col_ubi).value)
        if nom == "" and ubi == "":
            empty += 1
            if empty >= 80:
                break
            continue
        empty = 0
        if "ejemplo" in _key(nom):
            continue
        pares.add((_key(nom), _key(ubi)))
    return pares

# ============== UTILIDADES DGA ===============
REF_TOKENS = {"100", "-", "120", "350", "2500", "50", "65", "35", "720"}

def es_fila_referencia(vals):
    tokens = [str(v).strip().lower() for v in vals if v not in (None, "")]
    if not tokens:
        return True
    ref_hits = sum(t in REF_TOKENS for t in tokens)
    non_ref = len(tokens) - ref_hits
    return ref_hits >= 6 and non_ref <= 2

def leer_bloque_dga(ws):
    start_col = column_index_from_string(SRC_START_COL)
    end_col = column_index_from_string(SRC_LAST_COL)
    headers = [ws.cell(SRC_FIRST_DATA_ROW - 1, c).value for c in range(start_col, end_col + 1)]
    headers = [(_norm_str(h) or f"Col{i+1}") for i, h in enumerate(headers)]

    registros = []
    r = SRC_FIRST_DATA_ROW
    max_r = ws.max_row
    while r <= max_r:
        row_vals = [ws.cell(r, c).value for c in range(start_col, end_col + 1)]
        if all(_norm_str(v) in ("", "0") for v in row_vals):
            break
        if es_fila_referencia(row_vals):
            r += 1
            continue
        registros.append(row_vals)
        r += 1
    return headers, registros

# ============== PIPELINE =====================
def build_maestro() -> pd.DataFrame:
    """Lee TODAS las plantas y construye el maestro completo."""
    all_rows = []
    headers_global = None

    for folder in sorted(BASE.glob("* - Captura de datos")):
        archivos = [f for f in folder.glob("*.xlsm") if "transfor" in f.name.lower()]
        if not archivos:
            print(f"⚠️ {folder.name}: no hay archivo de transformadores, se omite.")
            continue

        xlsm = archivos[0]
        planta = folder.name.split(" - ")[0]
        print(f"📄 Procesando: {planta} ({xlsm.name})")

        wb = load_workbook(xlsm, data_only=True, read_only=True, keep_links=False)
        indice_pares = leer_pares_indice_wb(wb)
        print(f"📑 Índice: {len(indice_pares)} transformadores válidos")

        for ws in wb.worksheets:
            if hoja_ruidosa(ws.title):
                continue
            nom, ubi = read_name_loc_from_sheet(ws)
            if not nom or not ubi:
                continue
            if (_key(nom), _key(ubi)) not in indice_pares:
                continue

            headers, regs = leer_bloque_dga(ws)
            if not regs:
                continue
            if headers_global is None:
                headers_global = headers

            for fila in regs:
                row = {"Planta": planta, "Transformador": _norm_str(nom), "Ubicacion": _norm_str(ubi)}
                row.update({h: v for h, v in zip(headers, fila)})
                all_rows.append(row)

    if not all_rows:
        print("⚠️ No se obtuvieron registros válidos.")
        return pd.DataFrame()

    df = pd.DataFrame(all_rows)
    rename_map = {"Col1": "Compañía de Análisis", "Col2": "Fecha de Muestra", "Col3": "Fecha de Informe"}
    df = df.rename(columns=rename_map)

    drop_cols = [c for c in df.columns if "%" in str(c)]
    if drop_cols:
        df = df.drop(columns=drop_cols)

    if "Fecha de Muestra" in df.columns:
        df["_FechaM_dt"] = pd.to_datetime(df["Fecha de Muestra"], dayfirst=True, errors="coerce")
    else:
        df["_FechaM_dt"] = pd.NaT

    if "Fecha de Informe" in df.columns:
        df["_FechaI_dt"] = pd.to_datetime(df["Fecha de Informe"], dayfirst=True, errors="coerce")
    else:
        df["_FechaI_dt"] = pd.NaT

    mask_na = df["_FechaM_dt"].isna() & df["_FechaI_dt"].notna()
    df.loc[mask_na, "_FechaM_dt"] = df.loc[mask_na, "_FechaI_dt"]

    def fmt_dt(s):
        return s.dt.strftime("%d-%b-%y").where(s.notna(), other="NA")

    if "Fecha de Muestra" in df.columns:
        df["Fecha de Muestra"] = fmt_dt(df["_FechaM_dt"])
    if "Fecha de Informe" in df.columns:
        df["Fecha de Informe"] = fmt_dt(df["_FechaI_dt"])

    df = df.fillna("NA").replace("", "NA")
    return df

def calcular_ultimas(df: pd.DataFrame) -> pd.DataFrame:
    """Una fila por trafo (Planta/Transformador/Ubicacion) con la última 'Fecha de Muestra'."""
    if df.empty:
        return df

    if "_FechaM_dt" not in df.columns:
        base = pd.to_datetime(df.get("Fecha de Muestra"), dayfirst=True, errors="coerce")
        df = df.assign(_FechaM_dt=base)

    cols_key = ["Planta", "Transformador", "Ubicacion"]
    df_sorted = df.sort_values(cols_key + ["_FechaM_dt"], ascending=[True, True, True, True])
    ult = df_sorted.drop_duplicates(subset=cols_key, keep="last").copy()
    ult = ult.sort_values(cols_key).reset_index(drop=True)

    if "_FechaI_dt" in ult.columns:
        ult = ult.drop(columns=["_FechaI_dt"])
    if "_FechaM_dt" in ult.columns:
        ult = ult.drop(columns=["_FechaM_dt"])
    return ult

def escribir_excel(df_datos: pd.DataFrame, df_ult: pd.DataFrame):
    """Crea todas las hojas requeridas por la app."""
    with pd.ExcelWriter(OUT_FILE, engine="xlsxwriter") as writer:
        # ---- Datos ----
        df_datos.to_excel(writer, index=False, sheet_name="Datos")
        ws = writer.sheets["Datos"]
        nrows, ncols = df_datos.shape
        ws.add_table(0, 0, nrows, ncols - 1, {
            "columns": [{"header": c} for c in df_datos.columns],
            "name": "TablaDatos",
            "style": "Table Style Medium 9"
        })

        # ---- UltimaPorTrafo ----
        df_ult.to_excel(writer, index=False, sheet_name="UltimaPorTrafo")
        ws2 = writer.sheets["UltimaPorTrafo"]
        nrows2, ncols2 = df_ult.shape
        ws2.add_table(0, 0, nrows2, ncols2 - 1, {
            "columns": [{"header": c} for c in df_ult.columns],
            "name": "TablaUltimas",
            "style": "Table Style Medium 9"
        })

        # ---- Crear hojas vacías requeridas ----
        for hoja in ["Estados", "Diag_3Ratios", "Diag_IEC", "Diag_Duval"]:
            df_vacio = pd.DataFrame(columns=["Planta", "Transformador", "Ubicacion", f"Diagnóstico {hoja.split('_')[-1]}"])
            df_vacio.to_excel(writer, index=False, sheet_name=hoja)

# ================= MAIN ======================
def main():
    df = build_maestro()
    if df.empty:
        print("⚠️ No hubo datos para escribir.")
        return
    ult = calcular_ultimas(df)
    escribir_excel(df, ult)
    print(f"\n✅ Maestro + Hojas auxiliares generado: {OUT_FILE}")
    print(f"   Filas Datos: {len(df)} | Filas UltimaPorTrafo: {len(ult)}")

if __name__ == "__main__":
    main()