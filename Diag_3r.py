#!/usr/bin/env python3
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ================= RUTAS =================
BASE = Path("/Users/joseluisgiadanscastellanos/Library/CloudStorage/OneDrive-CEMEX/OneDrive_Cemex")
OUT_FILE = BASE / "Codigos/out/trafos_maestro_tabla.xlsx"

# ================= FUNCIONES =================
def _safe(a, b):
    """División segura para evitar /0"""
    try:
        a = float(a); b = float(b)
        if b == 0: return float("inf") if a > 0 else 0.0
        return a / b
    except:
        return 0.0

def diag_3ratios(row):
    ch4   = row.get("CH4", 0)
    c2h4  = row.get("C2H4", 0)
    c2h6  = row.get("C2H6", 0)
    c2h2  = row.get("C2H2", 0)
    h2    = row.get("H2", 0)

    R1 = _safe(c2h2, c2h4)   # C2H2/C2H4
    R2 = _safe(ch4,  h2)     # CH4/H2
    R3 = _safe(c2h4, c2h6)   # C2H4/C2H6

    if R1 > 3 and R3 > 3:
        diag = "D2 (Arcing)"
    elif R1 > 1 and R3 > 1:
        diag = "DT (Discharge + Thermal)"
    elif R3 > 1 and R2 < 1:
        diag = "T3 (>700°C)"
    elif 0.5 < R3 <= 1.0:
        diag = "T2 (300–700°C)"
    else:
        diag = "T1 (<300°C)"
    return round(R1,2), round(R2,2), round(R3,2), diag

# ================= MAIN =================
def main():
    if not OUT_FILE.exists():
        print(f"❌ No encuentro el archivo: {OUT_FILE}")
        return

    # Leer hoja UltimaPorTrafo
    df = pd.read_excel(OUT_FILE, sheet_name="UltimaPorTrafo")

    # Asegurar columnas
    needed = ["CH4","C2H4","C2H6","C2H2","H2"]
    for col in needed:
        if col not in df.columns:
            print(f"❌ Falta columna {col} en UltimaPorTrafo")
            return

    # Calcular ratios y diagnóstico
    results = []
    for _, row in df.iterrows():
        R1, R2, R3, diag = diag_3ratios(row)
        results.append({
            "Planta": row["Planta"],
            "Transformador": row["Transformador"],
            "Ubicacion": row["Ubicacion"],
            "Fecha de Muestra": row["Fecha de Muestra"],
            "R1 (C2H2/C2H4)": R1,
            "R2 (CH4/H2)": R2,
            "R3 (C2H4/C2H6)": R3,
            "Diagnóstico 3 Ratios": diag
        })
    out_df = pd.DataFrame(results)

    # Guardar como tabla en hoja nueva
    with pd.ExcelWriter(OUT_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        out_df.to_excel(writer, index=False, sheet_name="Diag_3Ratios")

    # Volvemos a abrir con openpyxl para añadir formato
    wb = load_workbook(OUT_FILE)
    ws = wb["Diag_3Ratios"]

    # ----- Convertir en tabla estilo Excel -----
    nrows, ncols = out_df.shape
    from openpyxl.worksheet.table import Table, TableStyleInfo
    table = Table(displayName="Tabla3Ratios", ref=f"A1:H{nrows+1}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    # ----- Colorear columna de diagnóstico -----
    col_diag = None
    for i, cell in enumerate(ws[1], start=1):
        if cell.value == "Diagnóstico 3 Ratios":
            col_diag = i
            break

    if col_diag:
        for r in range(2, ws.max_row+1):
            val = str(ws.cell(r, col_diag).value or "")
            if "T1" in val:
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Verde
            elif "T2" in val:
                fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Amarillo
            elif any(x in val for x in ["T3","D2","DT"]):
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Rojo
            else:
                fill = None
            if fill:
                ws.cell(r, col_diag).fill = fill

    wb.save(OUT_FILE)
    print(f"✅ Hoja 'Diag_3Ratios' añadida como tabla con colores a {OUT_FILE}")

if __name__ == "__main__":
    main()