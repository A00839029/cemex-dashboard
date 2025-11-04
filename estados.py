#!/usr/bin/env python3
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill

# ================= RUTAS =================
BASE = Path("/Users/joseluisgiadanscastellanos/Library/CloudStorage/OneDrive-CEMEX/OneDrive_Cemex")
OUT_FILE = BASE / "Codigos/out/trafos_maestro_tabla.xlsx"

# ================= LIMITES IEEE C57.104 ================
LIMITS = {
    "CH4": [(100, "Normal"), (1000, "Preocupante"), (float("inf"), "Crítico")],
    "C2H4": [(50, "Normal"), (500, "Preocupante"), (float("inf"), "Crítico")],
    "C2H2": [(5, "Normal"), (35, "Preocupante"), (float("inf"), "Crítico")],
    "TDGC": [(720, "Normal"), (1920, "Preocupante"), (float("inf"), "Crítico")],
}

# ================= FUNCIONES =================
def clasificar(valor, gas):
    try:
        v = float(valor)
    except:
        return "Normal"
    for lim, estado in LIMITS[gas]:
        if v <= lim:
            return estado
    return "Normal"

def estado_global(row):
    estados = []
    for gas in ["CH4", "C2H4", "C2H2", "TDGC"]:
        if gas in row:
            estados.append(clasificar(row[gas], gas))
    if "Crítico" in estados:
        return "Crítico"
    if "Preocupante" in estados:
        return "Preocupante"
    return "Normal"

# ================= MAIN =================
def main():
    if not OUT_FILE.exists():
        print(f"❌ No encuentro el archivo: {OUT_FILE}")
        return

    df = pd.read_excel(OUT_FILE, sheet_name="UltimaPorTrafo")

    # Buscar TDGC
    df = df.rename(columns=lambda x: x.strip())
    if "ppm" in df.columns:
        df = df.rename(columns={"ppm": "TDGC"})
    elif "TDCG" in df.columns:
        df = df.rename(columns={"TDCG": "TDGC"})

    # Clasificar
    df["Diagnóstico IEEE"] = df.apply(estado_global, axis=1)

    cols = ["Planta", "Transformador", "Ubicacion", "Fecha de Muestra", "Diagnóstico IEEE"]
    estados_df = df[cols].copy()

    # ==== Guardar (sin borrar formato del libro)
    wb = load_workbook(OUT_FILE)
    if "Estados" not in wb.sheetnames:
        wb.create_sheet("Estados")
    ws = wb["Estados"]

    # Limpia el contenido, pero no borra la hoja
    for row in ws["A2": f"Z{ws.max_row}"]:
        for cell in row:
            cell.value = None

    # Copia los datos
    for j, col in enumerate(estados_df.columns, start=1):
        ws.cell(1, j).value = col
    for i, row in enumerate(estados_df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j).value = val

    # Aplica colores por diagnóstico
    colores = {"Normal": "C6EFCE", "Preocupante": "FFF2CC", "Crítico": "FFC7CE"}
    col_idx = estados_df.columns.get_loc("Diagnóstico IEEE") + 1
    for i in range(2, ws.max_row + 1):
        val = ws.cell(i, col_idx).value
        if val in colores:
            ws.cell(i, col_idx).fill = PatternFill(start_color=colores[val], end_color=colores[val], fill_type="solid")

    wb.save(OUT_FILE)
    print(f"✅ Hoja 'Estados' (Diagnóstico IEEE) actualizada en {OUT_FILE}")

if __name__ == "__main__":
    main()